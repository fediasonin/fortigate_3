import os
import pandas as pd
import json
import time
import datetime
import logging
import asyncio
import aiohttp
import api_client
from date_util.date_filters_sorter import create_night_filter, group_and_sort, write_sorted_data_to_excel, \
    create_weekend_filter, merge_monthly_data_from_folder, filter_month_by_time_intervals

from pdql_filteres import event_filters
from event_analyzer import dataparse
from openpyxl import load_workbook



def run_vpn(d1, d2):
    new_column_names = {
        "time": "Время",
        "src.ip": "Внешний адрес",
        "src.geo.country": "Страна",
        "assigned_src_ip": "Выделенный адрес",
        "subject.name": "УЗ",
        "text": "Описание"
    }

    output_dir = 'output'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    current_time = time.localtime()
    formatted_time = time.strftime('_%d-%m-%Y_%H-%M-%S', current_time)

    output_file_xlsx = os.path.join(output_dir, f'events{formatted_time}.xlsx')

    writer = pd.ExcelWriter(output_file_xlsx, engine='xlsxwriter')

    ip_whitelist = dataparse.parse_ip_file("config/filtered_addresses.txt")


    #json_list = dataparse.csv_to_json_list("input.csv", ip_whitelist)
    json_list = asyncio.run(get_json_list_vpn(d1, d2, ip_whitelist))

    df = dataparse.json_to_dataframe(json_list)
    df = df.drop(columns=['uuid'])
    new_column = ['time', 'src.ip', 'src.geo.country', 'assigned_src_ip', 'subject.name', 'text']
    df = df[new_column]
    dataparse.dataframe_to_excel(df, writer, 'Все события', new_column_names)



    writer._save()
    print(f"Saving the output to {output_file_xlsx}")

    print("\nDone!")


def run_vpn_csv():
    new_column_names = {
        "time": "Время",
        "src.ip": "Внешний адрес",
        "src.geo.country": "Страна",
        "assigned_src_ip": "Выделенный адрес",
        "subject.name": "УЗ",
        "text": "Описание"
    }

    output_dir = 'output'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    current_time = time.localtime()
    formatted_time = time.strftime('_%d-%m-%Y_%H-%M-%S', current_time)

    output_file_xlsx = os.path.join(output_dir, f'events{formatted_time}.xlsx')

    writer = pd.ExcelWriter(output_file_xlsx, engine='xlsxwriter')

    ip_whitelist = dataparse.parse_ip_file("config/filtered_addresses.txt")

    json_list = dataparse.csv_to_json_list("input.csv", ip_whitelist)
    #json_list = asyncio.run(get_json_list_vpn(d1, d2, ip_whitelist))

    df = dataparse.json_to_dataframe(json_list)
    df = df.drop(columns=['uuid'])
    new_column = ['time', 'src.ip', 'src.geo.country', 'assigned_src_ip', 'subject.name', 'text']
    df = df[new_column]
    dataparse.dataframe_to_excel(df, writer, 'Все события', new_column_names)

    writer._save()
    print(f"Saving the output to {output_file_xlsx}")

    print("\nDone!")



def run_vpn_to_excel_week(input_file: str, output_file: str):
    os.chdir('output')
    df = pd.read_excel(input_file, engine='openpyxl')
    selected_columns = ['Время', 'Внешний адрес', 'Страна', 'Выделенный адрес', 'УЗ', 'Описание']
    df = df[selected_columns]
    df['Время'] = pd.to_datetime(df['Время'])


    # Примеры использования глобальной функции для создания различных фильтров
    mon_tue_night_filter = create_night_filter(df, 0, 20, 1, 8)  # Понедельник 20:00 - Вторник 08:00
    tue_wed_night_filter = create_night_filter(df, 1, 20, 2, 8)  # Вторник 20:00 - Среда 08:00
    wed_thu_night_filter = create_night_filter(df, 2, 20, 3, 8)  # Среда 20:00 - Четверг 08:00
    thu_fri_night_filter = create_night_filter(df, 3, 20, 4, 8)  # Четверг 20:00 - Пятница 08:00

    # Фильтр для будних ночей (с понедельника 20:00 до пятницы 08:00)
    weekday_nights_filter = (
            create_night_filter(df, 0, 20, 1, 8) |
            create_night_filter(df, 1, 20, 2, 8) |
            create_night_filter(df, 2, 20, 3, 8) |
            create_night_filter(df, 3, 20, 4, 8)
    )

    # Фильтр для выходных (с пятницы 20:00 до понедельника 08:00)
    weekend_filter = create_weekend_filter(df, 20, 8)


    # Combine filters for "Все события" sheet
    all_events_df = df[weekday_nights_filter | weekend_filter]

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        # Save filtered data to 'Все события' sheet
        all_events_df.to_excel(writer, sheet_name='Все события', index=False)

        start_date = df['Время'].min().normalize()
        day = start_date.day
        month = start_date.month
        mon_tue_data = df[mon_tue_night_filter]
        tue_wed_data = df[tue_wed_night_filter]
        wed_thu_data = df[wed_thu_night_filter]
        thu_fri_data = df[thu_fri_night_filter]
        weekdays_data = df[weekday_nights_filter]
        weekends_data = df[weekend_filter]


        #сортировка данных
        sorted_mon_tue_adr = group_and_sort(mon_tue_data, ['Внешний адрес', 'Страна'])
        sorted_mon_tue_uz = group_and_sort(mon_tue_data, ['УЗ'])

        sorted_tue_wed_adr = group_and_sort(tue_wed_data, ['Внешний адрес', 'Страна'])
        sorted_tue_wed_uz = group_and_sort(tue_wed_data, ['УЗ'])

        sorted_wed_thu_adr = group_and_sort(wed_thu_data, ['Внешний адрес', 'Страна'])
        sorted_wed_thu_uz = group_and_sort(wed_thu_data, ['УЗ'])

        sorted_thu_fri_adr = group_and_sort(thu_fri_data, ['Внешний адрес', 'Страна'])
        sorted_thu_fri_uz = group_and_sort(thu_fri_data, ['УЗ'])

        sorted_weekdays_adr = group_and_sort(weekdays_data, ['Внешний адрес', 'Страна'])
        sorted_weekdays_uz = group_and_sort(weekdays_data, ['УЗ'])

        sorted_weekends_adr = group_and_sort(weekends_data, ['Внешний адрес', 'Страна'])
        sorted_weekends_uz = group_and_sort(weekends_data, ['УЗ'])


        #вывод отсортированных данных
        # Запись данных по дням
        write_sorted_data_to_excel(writer, sorted_mon_tue_adr, sorted_mon_tue_uz,f'по дням {day}.{month} - {day + 7}.{month}', 1, 0)
        write_sorted_data_to_excel(writer, sorted_tue_wed_adr, sorted_tue_wed_uz,f'по дням {day}.{month} - {day + 7}.{month}', 1, 9)
        write_sorted_data_to_excel(writer, sorted_wed_thu_adr, sorted_wed_thu_uz,f'по дням {day}.{month} - {day + 7}.{month}', 1, 18)
        write_sorted_data_to_excel(writer, sorted_thu_fri_adr, sorted_thu_fri_uz,f'по дням {day}.{month} - {day + 7}.{month}', 1, 27)
        write_sorted_data_to_excel(writer, sorted_weekends_adr, sorted_weekends_uz,f'по дням {day}.{month} - {day + 7}.{month}', 1, 36)

        # Запись данных по будням и выходным
        write_sorted_data_to_excel(writer, sorted_weekdays_adr, sorted_weekdays_uz,f'Будни и Выходные {day}.{month} - {day + 7}.{month}', 1, 0)
        write_sorted_data_to_excel(writer, sorted_weekends_adr, sorted_weekends_uz,f'Будни и Выходные {day}.{month} - {day + 7}.{month}', 1, 9)



    # Load the workbook to modify 'Все события' sheet
    wb = load_workbook(output_file)
    ws_all_events = wb['Все события']

    # Automatically adjust the width of all columns in 'Все события'
    for col in ws_all_events.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws_all_events.column_dimensions[col[0].column_letter].width = adjusted_width

    for sheet_name in wb.sheetnames:
        if sheet_name != 'Все события':
            ws = wb[sheet_name]

            if sheet_name == f'по дням {day}.{month} - {day + 7}.{month}':
                day_names = [
                    f'Пн-Вт {day}.{month} - {day + 1}.{month}',
                    f'Вт-Ср {day + 1}.{month} - {day + 2}.{month}',
                    f'Ср-Чт {day + 2}.{month} - {day + 3}.{month}',
                    f'Чт-Пт {day + 3}.{month} - {day + 4}.{month}',
                    f'Выходные {day + 4}.{month} - {day + 7}.{month}'
                ]
                for i, day_name in enumerate(day_names):
                    ws.cell(row=1, column=i * 9 + 1, value=day_name)

            if sheet_name == f'Будни и Выходные {day}.{month} - {day + 7}.{month}':
                ws.cell(row=1, column=1, value=f'Будни {day}.{month} - {day + 4}.{month}')
                ws.cell(row=1, column=10, value=f'Выходные {day + 4}.{month} - {day + 7}.{month}')

            # Automatically adjust the width of all columns
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"Данные успешно сохранены и модифицированы в файл '{output_file}'")



def run_vpn_month(output_file, year, month):
    os.chdir('output_vpn')
    merge_monthly_data_from_folder('Время', year, month, output_file)
    # Пример использования функции
    filter_month_by_time_intervals(output_file, output_file)





async def get_json_list_vpn(date1, date2, ip_list):
    # Загружаем учетные данные из файла
    with open('config/credentials.json', 'r') as file:
        creds = json.load(file)

    ROOT_URL_API = creds["url_root_api"]
    USERNAME = creds["username"]
    PASSWORD = creds["password"]
    CLIENT_SECRET = creds["secret"]

    # Получаем токен доступа
    bearer_token = api_client.get_bearer_token(
        root_url_api=ROOT_URL_API,
        username=USERNAME,
        password=PASSWORD,
        client_secret=CLIENT_SECRET
    )

    # Устанавливаем уровень логирования на DEBUG
    logging.basicConfig(level=logging.DEBUG)

    time_from = int(time.mktime(time.strptime(date1, '%Y-%m-%d %H:%M:%S')))
    time_to = int(time.mktime(time.strptime(date2, '%Y-%m-%d %H:%M:%S')))

    events_buffer = []

    logging.debug(f"Initial time_from: {time_from}, time_to: {time_to}")

    while True:
        # Запрашиваем события через API
        events, total_count, last_incident_time = api_client.get_events_by_filter(
            root_url_api=ROOT_URL_API,
            access_token=bearer_token,
            filter=event_filters.vpn_mosreg_attacks,
            time_from=time_from,
            time_to=time_to
        )

        logging.debug(f"Fetched {total_count} events. Last incident time: {last_incident_time}")

        events_buffer += events

        # Если количество событий меньше или равно 10000, выходим из цикла
        if total_count <= 10000:
            logging.debug("Total count is 10000 or less, breaking the loop.")
            break

        # Обновляем time_to для следующего запроса
        time_to = int(datetime.datetime.strptime(last_incident_time, "%Y-%m-%dT%H:%M:%S.%f0Z").timestamp()) + 15000
        logging.debug(f"Updated time_to: {time_to}")

    # Убираем дублирующиеся события на основе UUID
    unique_events = {event['uuid']: event for event in events_buffer}
    total_events = list(unique_events.values())

    logging.debug(f"Total unique events: {len(total_events)}")

    # Обработка событий: фильтрация по IP и определение страны для IP-адресов
    count = 0
    ip_cache = {}

    async with aiohttp.ClientSession() as session:
        tasks = []

        for event in total_events:
            src_ip = event.get('src.ip')
            if dataparse.ip_in_list(src_ip, ip_list):
                total_events.remove(event)
                logging.debug(f"Event with src_ip {src_ip} removed from total_events.")
            elif src_ip:
                if src_ip not in ip_cache:
                    task = asyncio.create_task(api_client.get_country_by_ip(session, src_ip))
                    tasks.append(task)
                    ip_cache[src_ip] = await task
                    count += 1
                event['src.geo.country'] = ip_cache[src_ip]

        # Дожидаемся завершения всех асинхронных задач
        await asyncio.gather(*tasks)

    logging.debug(f"Final number of total events: {len(total_events)}")
    logging.info(f"{count} unique IP addresses were processed.")

    return total_events


