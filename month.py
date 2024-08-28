import pandas as pd
import os

from openpyxl.reader.excel import load_workbook

from date_util.date_filters_sorter import create_night_filter, group_and_sort, write_sorted_data_to_excel, \
    create_weekend_filter


def merge_monthly_data_from_folder(date_column, target_year, target_month, output_file):
    dataframes = []

    # Указываем папку относительно текущей директории

    # Получаем список всех файлов в папке
    files = [os.path.join(f) for f in os.listdir() if f.endswith('.xlsx')]

    for file in files:
        # Чтение Excel файла
        df = pd.read_excel(file)

        # Конвертация колонки с датой в формат datetime
        df[date_column] = pd.to_datetime(df[date_column])

        # Фильтрация данных по указанному месяцу и году
        filtered_df = df[(df[date_column].dt.year == target_year) &
                         (df[date_column].dt.month == target_month)]

        # Добавление отфильтрованного DataFrame в список
        dataframes.append(filtered_df)

    # Объединение всех отфильтрованных данных
    merged_df = pd.concat(dataframes)

    # Сортировка по времени
    merged_df = merged_df.sort_values(by=date_column)

    # Сохранение результата в один Excel файл
    merged_df.to_excel(output_file, index=False)


def filter_data_by_time_intervals(input_file: str, output_file: str):
    df = pd.read_excel(input_file, engine='openpyxl')

    # Конвертация колонки с временем в формат datetime
    df['Время'] = pd.to_datetime(df['Время'])

    # Фильтр для будних ночей (с понедельника 20:00 до пятницы 08:00)
    weekday_nights_filter = (
            create_night_filter(df, 0, 20, 1, 8) |
            create_night_filter(df, 1, 20, 2, 8) |
            create_night_filter(df, 2, 20, 3, 8) |
            create_night_filter(df, 3, 20, 4, 8)
    )

    # Фильтр для выходных (с пятницы 20:00 до понедельника 08:00)
    weekend_filter = create_weekend_filter(df, 20, 8)

    # Применение фильтров к данным
    filtered_df = df[weekday_nights_filter | weekend_filter]

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        # Сохранение отфильтрованных данных в новый Excel файл
        filtered_df.to_excel(writer, sheet_name='Все события', index=False)

        start_date = df['Время'].min().normalize()
        start_day = start_date.day
        month = start_date.month
        start_date = df['Время'].max().normalize()
        end_day = start_date.day
        weekdays_data = df[weekday_nights_filter]
        weekends_data = df[weekend_filter]

        sorted_weekdays_adr = group_and_sort(weekdays_data, ['Внешний адрес', 'Страна'])
        sorted_weekdays_uz = group_and_sort(weekdays_data, ['УЗ'])

        sorted_weekends_adr = group_and_sort(weekends_data, ['Внешний адрес', 'Страна'])
        sorted_weekends_uz = group_and_sort(weekends_data, ['УЗ'])

        # Запись данных по будням и выходным
        write_sorted_data_to_excel(writer, sorted_weekdays_adr, sorted_weekdays_uz,
                                   f'Будни и Выходные {start_day}.{month} - {end_day}.{month}', 1, 0)
        write_sorted_data_to_excel(writer, sorted_weekends_adr, sorted_weekends_uz,
                                   f'Будни и Выходные {start_day}.{month} - {end_day}.{month}', 1, 9)

    # Load the workbook to modify 'Все события' sheet
    wb = load_workbook(output_file)
    ws_all_events = wb['Все события']
    # Automatically adjust the width of all columns in 'Все события'
    for col in ws_all_events.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws_all_events.column_dimensions[col[0].column_letter].width = adjusted_width

    for sheet_name in wb.sheetnames:
        if sheet_name != 'Все события':
            ws = wb[sheet_name]

            if sheet_name == f'Будни и Выходные {start_day}.{month} - {end_day}.{month}':
                ws.cell(row=1, column=1, value=f'Будни')
                ws.cell(row=1, column=10, value=f'Выходные')

            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
    wb.save(output_file)


def run_vpn_month(output_file, year, month):
    os.chdir('output_vpn')
    merge_monthly_data_from_folder('Время', year, month, 'august_2024_sort.xlsx')
    # Пример использования функции
    filter_data_by_time_intervals('august_2024_sort.xlsx', output_file)




out = 'august_2024_sorted.xlsx'
y = 2024
m = 8
run_vpn_month(out, y, m)